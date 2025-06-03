import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

class ExcelUpdater:
    def __init__(self, filename="cloud_updates.xlsx"):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        self.headers = [
            "Provider", "Title", "URL", "Date Posted", "Description", "Links",
            "AWS Product", "Azure Products", "Azure Categories", "Azure Status", "Azure Update Type"
        ]
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        if os.path.exists(self.filename):
            try:
                self.workbook = load_workbook(self.filename)
                if "Updates" in self.workbook.sheetnames:
                    self.sheet = self.workbook["Updates"]
                else:
                    if len(self.workbook.sheetnames) == 1 and self.workbook.active.max_row <= 1: 
                        self.sheet = self.workbook.active
                        self.sheet.title = "Updates"
                    else: 
                        print(f"Sheet 'Updates' not found in {self.filename}. Creating it.")
                        self.sheet = self.workbook.create_sheet("Updates")
                
                if self.sheet.max_row > 0:
                    current_headers = [cell.value for cell in self.sheet[1]]
                    if len(current_headers) > 3 and "Title" in current_headers : 
                        self.headers = current_headers
                        print(f"Loaded existing workbook '{self.filename}' and sheet '{self.sheet.title}' with headers: {self.headers}")
                    else: 
                        print(f"Sheet '{self.sheet.title}' in '{self.filename}' has content but headers are missing/invalid. Appending standard headers.")
                        is_first_row_empty = True
                        if self.sheet.max_row >=1:
                            first_row_values = [cell.value for cell in self.sheet[1]]
                            if any(first_row_values):
                                is_first_row_empty = False

                        if self.sheet.max_row == 0 or is_first_row_empty:
                             self.sheet.append(self.headers)
                             print("Appended headers to existing but empty/headerless sheet.")
                        else:
                            print(f"Warning: Sheet '{self.sheet.title}' has existing data but unrecognized headers. Standard headers will be used for mapping, but not re-written to avoid data loss.")
                else: 
                    self.sheet.append(self.headers)
                    print(f"Loaded existing workbook '{self.filename}'. Sheet '{self.sheet.title}' was empty. Added headers.")

            except InvalidFileException:
                print(f"Error: File '{self.filename}' is not a valid Excel file or is corrupted. Creating a new workbook.")
                self._create_new_workbook()
            except Exception as e:
                print(f"An unexpected error occurred while loading workbook: {e}. Creating a new workbook.")
                self._create_new_workbook()
        else:
            self._create_new_workbook()

    def _create_new_workbook(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Updates"
        self.sheet.append(self.headers)
        print(f"Created new workbook '{self.filename}' and sheet 'Updates' with headers.")

    def add_update(self, data: dict):
        if not self.sheet:
            print("Error: Worksheet not initialized.")
            return

        row_to_add = []
        for header in self.headers:
            value = None
            if header == "Provider": value = data.get('provider')
            elif header == "Title": value = data.get('title')
            elif header == "URL": value = data.get('url')
            elif header == "Date Posted": value = data.get('date_posted')
            elif header == "Description": value = data.get('description')
            elif header == "Links": value = data.get('links')
            elif header == "AWS Product": value = data.get('product') if data.get('provider') == 'AWS' else None
            # Handle Azure metadata with priority from RSS feed if available
            elif header == "Azure Products":
                if data.get('provider') == 'Azure':
                    # Use product_list from data, which is already processed to include RSS data
                    value = data.get('product_list')
                else:
                    value = None
            elif header == "Azure Categories":
                if data.get('provider') == 'Azure':
                    value = data.get('categories')
                else:
                    value = None
            elif header == "Azure Status":
                if data.get('provider') == 'Azure':
                    value = data.get('status')
                else:
                    value = None
            elif header == "Azure Update Type":
                if data.get('provider') == 'Azure':
                    value = data.get('update_type')
                else:
                    value = None
            else: 
                value = data.get(header) 
            row_to_add.append(value if value is not None else "") 
        self.sheet.append(row_to_add)

    def save_workbook(self):
        if not self.workbook:
            print("Error: Workbook not initialized.")
            return
        try:
            self.workbook.save(self.filename)
            print(f"Workbook saved to {self.filename}")
        except Exception as e:
            print(f"Error saving workbook: {e}")

if __name__ == '__main__':
    TEST_FILENAME = "test_cloud_updates.xlsx"
    if os.path.exists(TEST_FILENAME):
        os.remove(TEST_FILENAME)

    print("--- Test 1: Creating new workbook ---")
    writer = ExcelUpdater(TEST_FILENAME)
    
    aws_sample_data = {
        'provider': 'AWS', 'title': 'AWS Test Update 1', 'url': 'http://aws.example.com/1',
        'date_posted': '06/10/2024', 'description': 'Desc for AWS 1', 'links': 'http://link1,http://link2',
        'product': 'EC2' 
    }
    azure_sample_data = {
        'provider': 'Azure', 'title': 'Azure Test Update 1', 'url': 'http://azure.example.com/1',
        'date_posted': '06/11/2024', 'description': 'Desc for Azure 1', 'links': 'http://link3',
        'product_list': 'VMs, Storage', 
        'categories': 'Compute, Storage',
        'status': 'Launched',
        'update_type': 'Features'
    }
    
    writer.add_update(aws_sample_data)
    writer.add_update(azure_sample_data)
    writer.save_workbook()
    print(f"Current headers after Test 1: {writer.headers}")

    print("\n--- Test 2: Loading existing workbook and adding more data ---")
    writer2 = ExcelUpdater(TEST_FILENAME)
    print(f"Headers loaded in writer2: {writer2.headers}") 

    another_aws_data = {
        'provider': 'AWS', 'title': 'AWS Test Update 2', 'url': 'http://aws.example.com/2',
        'date_posted': '06/12/2024', 'description': 'More AWS stuff', 'links': '',
        'product': 'S3'
    }
    writer2.add_update(another_aws_data)
    
    azure_custom_data = {
        'provider': 'Azure', 'title': 'Azure Test Update 2 with custom field', 'url': 'http://azure.example.com/2',
        'date_posted': '06/13/2024', 'description': 'Desc for Azure 2', 'links': 'http://link4',
        'product_list': 'AI', 'categories': 'AI + Machine Learning',
        'status': 'Preview', 'update_type': 'Services', 'CustomInfo': 'TestValue' 
    }
    if "CustomInfo" not in writer2.headers:
         pass 
    writer2.add_update(azure_custom_data) 
    writer2.save_workbook()

    print("\n--- Test 3: Verifying content (first few rows) ---")
    try:
        wb_verify = load_workbook(TEST_FILENAME)
        sheet_verify = wb_verify["Updates"]
        print("Headers from file:", [cell.value for cell in sheet_verify[1]])
        print("Row 2:", [cell.value for cell in sheet_verify[2]])
        print("Row 3:", [cell.value for cell in sheet_verify[3]])
        print("Row 4:", [cell.value for cell in sheet_verify[4]])
        if sheet_verify.max_row > 4: 
            print("Row 5:", [cell.value for cell in sheet_verify[5]])
    except Exception as e:
        print(f"Error verifying file: {e}")
    
    print("\n--- Test 4: Handling of corrupted or non-Excel file ---")
    CORRUPTED_FILENAME = "corrupted_test_cloud_updates.xlsx"
    with open(CORRUPTED_FILENAME, "w") as f:
        f.write("This is not an excel file.")
    writer_corrupt = ExcelUpdater(CORRUPTED_FILENAME) 
    writer_corrupt.add_update(aws_sample_data)
    writer_corrupt.save_workbook()
    if os.path.exists(CORRUPTED_FILENAME): 
        try:
            load_workbook(CORRUPTED_FILENAME)
            print(f"File '{CORRUPTED_FILENAME}' was successfully created/overwritten as a valid Excel file.")
        except Exception as e:
            print(f"File '{CORRUPTED_FILENAME}' is still corrupted or not a valid Excel file: {e}")

    print(f"\nNote: Test files '{TEST_FILENAME}' and '{CORRUPTED_FILENAME}' were created/updated.")
    print("Please inspect them manually if needed. Uncomment cleanup lines to auto-delete.")
# Removed the erroneous ``` line that was here
